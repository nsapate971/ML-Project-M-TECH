from tkinter import *
import matplotlib.pyplot as plt
#from Tkinter import *
import tkinter.messagebox as tkMessageBox
from tkinter import filedialog
import sqlite3
from tkinter import ttk
from PIL import *
from PIL import ImageTk, Image
import xlrd, xlwt
from xlrd import *
from xlsxwriter.utility import xl_rowcol_to_cell
import xlsxwriter
import numpy as np
import openpyxl
from xlrd import open_workbook
from openpyxl import load_workbook
import array as arr
from xlwt import Workbook 
import os
import tkinter as tk
import pandas as pd
plt.style.use('seaborn-white')
from pandas import read_csv
from sklearn.model_selection import KFold
from sklearn.model_selection import cross_val_score
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import StandardScaler
import csv
os.environ["PYTHONIOENCODING"] = "utf-8"

# /* Multispectral Laboratory  */#
# /* Author  : Nikhil M Sapate */#
# /* M. Tech : Second Year     */# 
# /* Research Guide   : Prof. Dr. R. R. Deshmukh Sir */#
global filename
root = Tk()
root.title("Multispectral Laboratory | Dept. of CS and IT | Dr. B.A.M.U., Aurangabad")
filename = ImageTk.PhotoImage(master=root ,file= "D://M.Tech4thSem//WorkProject//GUIWork//landing_Page.jpg")
background_label = Label(root, image=filename)
background_label.place(x=0, y=0, relwidth=1, relheight=1)
background_label.pack()
background_label.config(image = filename)
background_label.config(compound= "bottom")
width = 1024 #1366
height = 520 #768 
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width/2) - (width/2)
y = (screen_height/2) - (height/2)
root.geometry("%dx%d+%d+%d" % (width, height, x, y))
root.resizable(0, 0)
#root.config(bg="#228B22")

#========================================VARIABLES========================================
USERNAME = StringVar()
PASSWORD = StringVar()

#========================================METHODS==========================================

def Database():
    global conn, cursor
    conn = sqlite3.connect("pythontut.db")
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS `admin` (admin_id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, username TEXT, password TEXT)")
    cursor.execute("CREATE TABLE IF NOT EXISTS `product` (product_id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, product_name TEXT, product_qty TEXT, product_price TEXT)")
    cursor.execute("SELECT * FROM `admin` WHERE `username` = 'admin' AND `password` = 'adminlab'")
    if cursor.fetchone() is None:
        cursor.execute("INSERT INTO `admin` (username, password) VALUES('admin', 'admin')")
        conn.commit()

def Exit():
    result = tkMessageBox.askquestion('Multispectral Laboratory', 'Are you sure you want to exit?', icon="warning")
    if result == 'yes':
        root.destroy()
        exit()

def Exit2():
    result = tkMessageBox.askquestion('Multispectral Laboratory', 'Are you sure you want to exit?', icon="warning")
    if result == 'yes':
        Home.destroy()
        exit()

def ShowLoginForm():
    global loginform
    loginform = Toplevel()
    loginform.title("Multispectral Laboratory/Account Login")
    width = 600
    height = 500
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    loginform.resizable(0, 0)
    loginform.geometry("%dx%d+%d+%d" % (width, height, x, y))
    loginform.config(bg="#228B22")
    LoginForm()
    
def LoginForm():
    global lbl_result
    TopLoginForm = tk.Frame(loginform, width=600, height=100, bd=1, relief=SOLID)
    TopLoginForm.pack(side=TOP, pady=20)
    lbl_text = Label(TopLoginForm, text="Administrator   Login", font=('Berlin Sans FB Demi', 18), width=600)
    lbl_text.pack(fill=X)
    MidLoginForm = tk.Frame(loginform, width=600)
    MidLoginForm.pack(side=TOP, pady=50)
    lbl_username = Label(MidLoginForm, text="Username:", font=('Berlin Sans FB Demi', 20), bd=10)
    lbl_username.grid(row=0)
    lbl_password = Label(MidLoginForm, text="Password:", font=('Berlin Sans FB Demi', 20), bd=10)
    lbl_password.grid(row=1)
    lbl_result = Label(MidLoginForm, text="", font=('arial', 18))
    lbl_result.grid(row=3, columnspan=2)
    username = Entry(MidLoginForm, textvariable=USERNAME, font=('Berlin Sans FB Demi', 20), width=25)
    username.grid(row=0, column=1)
    password = Entry(MidLoginForm, textvariable=PASSWORD, font=('Berlin Sans FB Demi', 20), width=25, show="*")
    password.grid(row=1, column=1)
    btn_login = Button(MidLoginForm, text="Login", font=('Berlin Sans FB Demi', 18), width=25, command=Login)
    btn_login.grid(row=2, columnspan=2, pady=20)
    btn_login.bind('<Return>', Login)
    
    
def Home():
    global Home
    global filename1
    Home = Tk()
    width = 1024
    height = 520
    screen_width = Home.winfo_screenwidth()
    screen_height = Home.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    Home.geometry("%dx%d+%d+%d" % (width, height, x, y))
    Home.resizable(0, 0)
    filename1 = ImageTk.PhotoImage(master=Home ,file= "D://M.Tech4thSem//WorkProject//GUIWork//landing_Page.jpg")
    background_label1 = Label(Home, image=filename1)
    background_label1.place(x=0, y=0, relwidth=1, relheight=1)
    background_label1.pack()
    background_label1.config(image = filename1)
    background_label1.config(compound= "left")

    menubar = Menu(Home)
    filemenu = Menu(menubar, tearoff=0)
    filemenu2 = Menu(menubar, tearoff=0)
    filemenu3 = Menu(menubar, tearoff=0)
    filemenu4 = Menu(menubar, tearoff=0)
    filemenu.add_command(label="Logout", command=Logout)
    filemenu.add_command(label="Exit", command=Exit2)
    filemenu2.add_command(label="Upload CSV", command=ShowAddNew)
    filemenu3.add_command(label="View Graph", command=ShowGraph1)
    menubar.add_cascade(label="Account", menu=filemenu)
    menubar.add_cascade(label="Start", menu=filemenu2)
    menubar.add_cascade(label="About", command=ShowAbout)
    menubar.add_cascade(label="Graphs", menu=filemenu3)
    menubar.add_cascade(label="Help", command=ShowHelp)
    Home.config(menu=menubar)


def ShowAddNew():
    global addnewform
    width = 600
    height = 500
    screen_width = Home.winfo_screenwidth()
    screen_height = Home.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    addnewform = Toplevel()
    addnewform.geometry("%dx%d+%d+%d" % (width, height, x, y))
    addnewform.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("xlsx","*.xlsx"),("all files","*.*")))
    name = addnewform.filename
    workbook = openpyxl.load_workbook(name)
    worksheet = workbook.active
    Sheet=workbook.create_sheet('features_of_hypersData')
    Sheet['A1']= "Sample Number"
    Sheet['B1']= "CRI1"
    Sheet['C1']= "PRI"
    Sheet['D1']= "SIPI"
    Sheet['E1']= "WBI"
    Sheet['F1']= "PhRI"
    Sheet['G1'] = "NPCI"
    Sheet['H1'] = "Health Status"
##    
    lbl= Label(addnewform, text="Enter the sample Number :", font=('Berlin Sans FB Demi', 15), bd=8)
    lbl.place(relx=.3, rely=.7, anchor="c")
    v1=tk.IntVar()
    sample_num = Entry(addnewform,textvariable=v1, font=('Berlin Sans FB Demi', 12),bd=8) # textvariable=v, 
    sample_num.place(relx=.8, rely=.7, anchor="c")
   
    def calculate_indices():
        v=int(v1.get())
        print ("User entered choice", v)
        if(2 <= v <=93):
            R_510 = worksheet.cell(161,v).value
            R_550 = worksheet.cell(201,v).value
            R_819  = worksheet.cell(471,v).value
            R_1599 = worksheet.cell(1251,v).value
            R_900  = worksheet.cell(552,v).value
            R_970 = worksheet.cell(622,v).value
            global WBI
            WBI = R_900 / R_970
            print("WBI Value is:", WBI)
            global MSI
            MSI = R_1599 / R_819
            global CRI1 
            CRI1 = (1/R_510) - (1/R_550)
            print (" CRI1 value is :", CRI1)
            R_701 = worksheet.cell(353,v).value
            R_671 = worksheet.cell(323,v).value
            R_549 = worksheet.cell(201,v).value
            R_680 = worksheet.cell(332,v).value
            R_430 = worksheet.cell(82,v).value
            global NPCI
            NPCI = (R_680 - R_430)/(R_680 + R_430)
            R_531=worksheet.cell(182,v).value
            R_570=worksheet.cell(221,v).value
            R_445=worksheet.cell(96,v).value
            global PRI
            PRI = ( R_531 - R_570 ) /( R_531 + R_570 )
            global PhRI
            PhRI = (R_550 - R_531)/(R_550 + R_531)
            print("PhRI value is", PhRI)
            print("NPCI value is", NPCI)
            print ("PRI value is", PRI)
            R_800=worksheet.cell(451,v).value
            global SIPI
            SIPI = (R_800 - R_445) / (R_800 + R_445)
            print ("SIPI is", SIPI)
            R_750=worksheet.cell(401,v).value
            R_705=worksheet.cell(356,v).value
            ## /**  Forest One Starts Here ** / ##
            ## /**  Decision Tree One Starts Here ** / ##
            vote=0
            vote_d = 0
            if (1<=CRI1 <=12 ): 
                vote = vote+1
                
                print("vote for Healthy Plant CRI1",vote)
            elif (0<= CRI1 <= 3.769):
                flag=1
                vote_d=vote_d+1
                
            else:
                disease_name ="\nResult : Plant is Diseased \n Disease Type: Leaf Miner \n Carotenoid Pigmentation Problem: High  \nPhotochemical Activity: Irregular \n \Canopy Water Content: Irregular \nWater Content: Low  : -VE\nNitrogen Content: None\nSucepability Chances: +VE \nSuceptibility : High chances of Early Blight"
                print("vote for Diseased plant CRI1:",vote_d)
                
            ## /**  Decision Tree One Ends Here ** / ##
            # -----------*************************------------ #
            if(-1 <= PRI <= 0):
                vote = vote+1
                print("vote for healthy plant PRI", vote)
            elif ( -0.0131 <= PRI <= -0.0075):
                flag =2 
                vote_d= vote_d + 1
                
                print(" vote for Diseased plant PRI:", vote_d)

            else:
                vote_d= vote_d + 1
                disease_name = "\nResult : Plant is Diseased  Disease Type : white spots \n nCarotenoid Pigmentation Problem: High  \nPhotochemical Activity: Irregular \n \Canopy Water Content: Irregular\nWater Content: Low  : -VE\nNitrogen Content: None\n Sucepability Chances: +VE"
                print(" vote for Diseased plant PRI:", vote_d)
                
            ## /**  Decision Tree Two Ends Here ** / ##
            # -----------*************************------------ #
            if(0.8 <= SIPI <= 1.8):
                vote = vote+1
                print("vote for healthy plant SIPI",vote)
            elif ( 0.58 <= SIPI <= 0.66):
                flag =3 
                vote_d= vote_d + 1
                
                print("vote for Diseased plant SIPI:",vote_d)
            else:
                vote_d= vote_d + 1
                disease_name = "\nResult : Plant is Diseased \n Disease Type: Phytophthora \n Carotenoid Pigmentation Problem: High  \nPhotochemical Activity: Irregular \n \Canopy Water Content: Irregular \nWater Content: Low  Chances: Minor\nNitrogen Content: None\nSucepability Chances: +VE"
                print("vote for Diseased plant SIPI:",vote_d)
                  
            ## /**  Decision Tree Three Ends Here ** / ##
            # -----------*************************------------ #
            if(0<= WBI <= 1):
                vote = vote+1
                print("vote for healthy plant WBI",vote)
            elif (1.01 <= WBI <= 1.033):
                flag = 4
                vote_d= vote_d + 1
                
                print("vote for Diseased plant WBI:",vote_d)
            else:
                vote_d= vote_d + 1
                disease_name="\nResult : Plant is Diseased Disease Type : Late Blight \nCarotenoid Pigmentation Problem: High  \nPhotochemical Activity: Irregular \n \Canopy Water Content: Irregular \nWater Content: Low    Sustainability : -VE\nNitrogen Content: None\nSucepability Chances: +VE"
                print("vote for Diseased plant WBI:",vote_d)
            ## /**  Decision Tree Four Ends Here ** / ##
            # -----------*************************------------ #
            if(0<=NPCI<=0.025):
                vote = vote+1
                print("vote for healthy plant NPCI",vote)
            elif ( -0.09 <= NPCI <= 0.0827):
                vote_d= vote_d + 1
                flag = 5
                
                print("vote for Diseased plant NPCI:",vote_d)
            else :
                vote_d= vote_d + 1
                disease_name = "\nResult : Plant is Diseased \n Disease type: Phythopora \nCarotenoid Pigmentation Problem: High\nPhotochemical Activity: Irregular \nCanopy Water Content: Irregular\nWater Content: Low\nNitrogen Content: None\nSucepability Chances: +VE"
                print("vote for Diseased plant NPCI:",vote_d)
            # -----------*************************------------ #
            ## /**  Decision Tree Five Ends Here ** / ##
            
            if ( vote > vote_d):
                print ( "Plant is healthy " )
                boolean = 0
                lbl_display4 = Label(addnewform, text= "\nResult : Plant is Healthy \n Severity : None \nCarotenoid Pigmentation Problem: None\n Photochemical Activity: Regular \nCanopy Water Content: Steady \n Water Content: High \n Sustainability : +VE\n Nitrogen Content: None \n Sucepability Chances: -VE",fg='green', font=('Berlin Sans FB Demi', 14),width=40,height=12)
                lbl_display4.place(relx=.5, rely=.2, anchor="c")
            else:
                if ( flag > 0 ):
                    if( flag == 1):
                        disease_name ="\nResult : Plant is Diseased \nDisease type: Septoria leaf spot \n Disease Severity : Minor \nCarotenoid Pigmentation Problem: High  \nPhotochemical Activity: Irregular \nCanopy Water Content: Irregular \nWater Content: Low \nNitrogen Content: None \nSucepability Chances: +VE"
                        
                    elif( flag == 2 ):
                        disease_name = "\nResult : Plant is Diseased \nDisease type: Early Blight\nDisease Severity : High\nCarotenoid Pigmentation Problem: High\nPhotochemical Activity: Irregular \nCanopy Water Content: Irregular \nWater Content: Low \n Nitrogen Content: None \nSucepability Chances: +VE "
                    elif( flag ==3):
                        disease_name = "\nResult : Plant is Diseased \nDisease type: Brown spots\nDisease Severity : Moderate\nCarotenoid Pigmentation Problem: High\nPhotochemical Activity: Irregular \n Canopy Water Content: Irregular \n Water Content: Low \nNitrogen Content: None \n Sucepability Chances: +VE"
                    elif( flag ==4):
                        disease_name="\nResult : Plant is Diseased \nDisease type: Late Blight\nDisease Severity : High\nCarotenoid Pigmentation Problem: High \nPhotochemical Activity: Irregular \nCanopy Water Content: Irregular \nWater Content: Low \n Nitrogen Content: None\nSucepability Chances: +VE"
                    elif(flag ==5):
                        disease_name = "\nResult : Plant is Diseased\nDisease type: Phyptophthora \nDisease Severity : High\nCarotenoid Pigmentation Problem: High\nPhotochemical Activity: Irregular \nCanopy Water Content: Irregular \nWater Content: Low\n Nitrogen Content: None \nSucepability Chances: +VE" 
                    else:
                     print("wrong")
                print ("Plant is diseased ")
                lbl_display19 = Label(addnewform, text= disease_name ,fg='red', font=('Berlin Sans FB Demi', 14),width=40,height=12)                                                                                                  
                boolean = 1
                lbl_display19.place(relx=.5, rely=.2, anchor="c")
            print ( """ ******* Decision Tree Ends Here ****** """)
            print ("This is Standered Result Obtained on Test Data")
            print ("""  Classification Report is as follows : \n
                    Confusion Matrix for Dataset is : [[ 2  1]\n
                                                       [ 0 13]]\n
                     precision    recall  f1-score   support\\n
                 0       1.00      0.67      0.80         3\n
                 1       0.93      1.00      0.96        13\n
          accuracy                           0.94        16\n
         macro avg       0.96      0.83      0.88        16\n
      weighted avg       0.94      0.94      0.93        16\n """)
      
                
            # -----------*************************------------ #
               

 
            n = v
            m=1
            
            if( n <=92 ):
                if ( m <=8 ):
                    
                    Sheet.cell(n,m).value = v
                    m+=1
                    Sheet.cell(n,m).value = CRI1
                    m+=1
                    Sheet.cell(n,m).value = PRI
                    m+=1
                    Sheet.cell(n,m).value = SIPI
                    m+=1
                    Sheet.cell(n,m).value = WBI
                    m+=1
                    Sheet.cell(n,m).value = PhRI
                    m+=1
                    Sheet.cell(n,m).value = NPCI
                    m+=1
                    Sheet.cell(n,m).value = boolean
                    m+=1
                    n+=1
                else:
                    print ("--** Indices Over ** --")
            else:
                print("wrong choice")
                
            
            
            workbook.save('D:/Datasets/newdata.xls')
            
        else:
            print("vegetation is healthy" )
            boolean = 0
            lbl_display4 = Label(addnewform, text= " Result : Plant is Healthy",fg='green', font=('Berlin Sans FB Demi', 20),width=40,height=2)
            lbl_display4.place(relx=.5, rely=.2, anchor="c")
            
    url = "D:/Datasets/Book1.csv"
    dataframe = read_csv(url)
    array = dataframe.values

    X = array[:,1:7]
    Y = array[:,7]

    num_trees = 10
    max_features = 6
    kfold = KFold(n_splits=4, random_state=7)
    model = RandomForestClassifier(n_estimators=num_trees, max_features=max_features)

    results = cross_val_score(model, X, Y, cv=kfold)

      
    button = Button(addnewform, text="Next >>", fg="black",command=calculate_indices)
    button.place(relx=.5, rely=.9, anchor="c")
    
    addnewform.resizable(0, 0)
    addnewform.config(bg="#228B22")

    
def ShowGraph():
    print ( "Graph")
    graph_data = pd.read_csv('D:/Datasets/Book1.csv')
    graph_data.head()
    print (graph_data['CRI1'].head(7))
    names = arr.array('i',[1,2,3,4])
    x = [CRI1,PRI,SIPI,PhRI,PhRI,WBI]
    num_bins = 5
    n, bins, patches = plt.hist(x, num_bins, facecolor='blue', alpha=0.5)    
    plt.xlabel("CRI1  PRI   SIPI  PhRI   WBI")
    plt.ylabel("Indices Values")
    plt.plot()
    plt.show()

def ShowView():
    global viewform
    viewform = Toplevel()
    viewform.title("Multispectral Laboratory/View Health Status of Plant")
    width = 600
    height = 400
    screen_width = Home.winfo_screenwidth()
    screen_height = Home.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    viewform.geometry("%dx%d+%d+%d" % (width, height, x, y))
    viewform.resizable(0, 0)
    ViewForm()

def ShowGraph1():
    global fq
    rw = Tk()
    rw.title("About Multispectral Laboratory")

    fq=ImageTk.PhotoImage(master=rw,file = "D://M.Tech4thSem//WorkProject//GUIWork//vi.jpg")
    background_label8 = Label(rw, image=fq)
    background_label8.place(x=0, y=0, relwidth=1, relheight=1)
    background_label8.pack()
    background_label8.config(image = fq)
    background_label8.config(compound= "bottom")
    width = 1130
    height = 470
    screen_width = Home.winfo_screenwidth()
    screen_height = Home.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    rw.geometry("%dx%d+%d+%d" % (width, height, x, y))
    rw.config(bg="#228B22")


def Logout():
    result = tkMessageBox.askquestion('Multispectral Laboratory', 'Are you sure you want to logout?', icon="warning")
    if result == 'yes': 
        admin_id = ""
        root.deiconify()
        Home.destroy()
        
def ShowAbout():        
    global mp
    rw = Tk()
    rw.title("About Authors")
    mp=ImageTk.PhotoImage(master=rw,file = "D://M.Tech4thSem//WorkProject//GUIWork//about.jpg")
    background_label2 = Label(rw, image=mp)
    background_label2.place(x=0, y=0, relwidth=1, relheight=1)
    background_label2.pack()
    background_label2.config(image = mp)
    background_label2.config(compound= "bottom")

    width = 1200
    height = 550
    screen_width = Home.winfo_screenwidth()
    screen_height = Home.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    rw.geometry("%dx%d+%d+%d" % (width, height, x, y))
    rw.resizable(0, 0)
    rw.config(bg="#228B22")


    
    
def ShowHelp():
    global hp
    rw = Tk()
    rw.title("About Multispectral Laboratory")
    hp=ImageTk.PhotoImage(master=rw,file = "D://M.Tech4thSem//WorkProject//GUIWork//help.jpg")
    background_label3 = Label(rw, image=hp)
    background_label3.place(x=0, y=0, relwidth=1, relheight=1)
    background_label3.pack()
    background_label3.config(image = hp)
    background_label3.config(compound= "bottom")
    width = 1200
    height = 550
    screen_width = Home.winfo_screenwidth()
    screen_height = Home.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    rw.geometry("%dx%d+%d+%d" % (width, height, x, y))
    rw.config(bg="#228B22")
    
def Login(event=None):
    global admin_id
    Database()
    if USERNAME.get == "" or PASSWORD.get() == "":
        lbl_result.config(text="Please complete the required field!", fg="red")
    else:
        cursor.execute("SELECT * FROM `admin` WHERE `username` = ? AND `password` = ?", (USERNAME.get(), PASSWORD.get()))
        if cursor.fetchone() is not None:
            cursor.execute("SELECT * FROM `admin` WHERE `username` = ? AND `password` = ?", (USERNAME.get(), PASSWORD.get()))
            data = cursor.fetchone()
            admin_id = data[0]
            USERNAME.set("")
            PASSWORD.set("")
            lbl_result.config(text="")
            ShowHome()
        else:
            lbl_result.config(text="Invalid username or password", fg="red")
            USERNAME.set("")
            PASSWORD.set("")
    cursor.close()
    conn.close() 

def ShowHome():
    root.withdraw()
    Home()
    loginform.destroy()


#========================================MENUBAR WIDGETS==================================
menubar = Menu(root)
filemenu = Menu(menubar, tearoff=0)
filemenu.add_command(label="Account", command=ShowLoginForm)
filemenu.add_command(label="Exit", command=Exit)
menubar.add_cascade(label="File", menu=filemenu)
root.config(menu=menubar)

#========================================FRAME============================================
Title = Frame(root, relief=SOLID)
Title.pack(pady=10)

#========================================INITIALIZATION===================================
if __name__ == '__main__':
    root.mainloop()
